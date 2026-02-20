# Copyright 2026 VentorTech OU
# License LGPL-3.0 or later (http://www.gnu.org/licenses/lgpl.html).

import base64
import io
from openpyxl import load_workbook

from datetime import date, datetime
from dateutil.relativedelta import relativedelta

from odoo import models, fields, api, _
from odoo.exceptions import UserError


class HrPayslipImportWizard(models.TransientModel):
    _name = "hr.payslip.import.wizard"
    _description = "Wizard: Import Payslip"

    import_file = fields.Binary(
        string="Original file",
        required=True,
        attachment=False,
        help="Upload XLs File",
    )

    date_from = fields.Date(
        string="From",
        required=True,
        default=lambda self: fields.Date.to_string(date.today().replace(day=1)),
    )

    date_to = fields.Date(
        string="To",
        required=True,
        default=lambda self: fields.Date.to_string(
            (datetime.now() + relativedelta(months=+1, day=1, days=-1)).date()
        ),
    )

    last_period = fields.Char(
        string="Last Period",
        compute="_compute_last_period",
    )

    message = fields.Text(
        default="",
    )

    alert = fields.Text(
        default="",
    )

    @api.depends("date_from")
    def _compute_last_period(self):
        last_period_value = ""
        last_batch = self.env["hr.payslip.run"].search(
            [], order="date_end desc", limit=1
        )
        if last_batch:
            last_period_value = "{} - {}".format(
                last_batch.date_start.strftime("%d/%m/%Y"),
                last_batch.date_end.strftime("%d/%m/%Y"),
            )
        self.last_period = last_period_value

    def _decode_file(self, file):
        try:
            decoded_file = base64.b64decode(file)
        except Exception as e:
            raise ValueError("Invalid base64 input: %s" % e)

        return io.BytesIO(decoded_file)

    def finish_import(self):
        return {
            "name": _("Import Payslips"),
            "type": "ir.actions.act_window",
            "res_model": "hr.payslip.import.wizard",
            "view_mode": "form",
            "view_type": "form",
            "res_id": self.id,
            "views": [(False, "form")],
            "target": "new",
        }

    def _get_existing_data(self, model, names, data_name):
        # check imported data for empty values and make sure
        # that imported data exists in the system
        not_empty_names = list(filter(None, names))
        not_empty_names = list(map(lambda name: name.strip(), not_empty_names))
        if len(not_empty_names) != len(names):
            self.alert += "File contains empty values in the {} data!\n".format(
                data_name
            )

        objs = self.env[model].search([("name", "in", not_empty_names)])
        dif = set(not_empty_names) - set(objs.mapped("name"))
        if dif:
            self.alert += (
                "{}s with the following names "
                "were not found in the system: \n{}\n".format(data_name, ", ".join(dif))
            )
        return objs

    def validate_employees(self, empl_names):
        employees = self._get_existing_data("hr.employee", empl_names, "Employee")
        employee_data = {empl.name: empl for empl in employees}
        return employee_data

    def validate_input_types(self, input_type_names):
        input_types = self._get_existing_data(
            "hr.payslip.input.type", input_type_names, "Input Type"
        )
        # sort objects in the order in which they are written in the file
        sorted_input_types = input_types.sorted(
            lambda x: input_type_names.index(x.name)
        )
        return sorted_input_types

    def create_batch(self):
        batch = self.env["hr.payslip.run"].create(
            {
                "name": "From {} to {}".format(
                    self.date_from.strftime("%d/%m/%Y"),
                    self.date_to.strftime("%d/%m/%Y"),
                ),
                "date_start": self.date_from,
                "date_end": self.date_to,
                "state": "01_ready",
            }
        )
        return batch

    def create_payslip(self, payslip_vals):
        batch = self.create_batch()
        created = 0
        for data in payslip_vals:
            for employee, input_types in data.items():
                non_zero_types = [
                    x
                    for x in input_types
                    if x[1] and not isinstance(x[1], str) and x[1] > 0
                ]
                input_lines = [
                    (
                        0,
                        0,
                        {
                            "input_type_id": input_type[0],
                            "amount": input_type[1],
                        },
                    )
                    for input_type in non_zero_types
                ]

                last_contract = employee.current_version_id

                payslip = self.env["hr.payslip"].create(
                    {
                        "name": "",
                        "payslip_run_id": batch.id,
                        "employee_id": employee.id,
                        "date_from": self.date_from,
                        "date_to": self.date_to,
                        "version_id": last_contract.id if last_contract else None,
                        "input_line_ids": input_lines,
                    }
                )
                # _compute_name() isn't run while creating object, run it explicitly
                payslip._compute_name()
                created += 1

        self.message += (
            "Payroll import completed successfully.\n"
            "{} payslip(s) were created (batch - {})".format(created, batch.name)
        )

    def import_payslip(self):
        self.ensure_one()

        # delete alert message if new file was uploaded in the same wizard
        if self.alert:
            self.alert = ""

        fileobj = self._decode_file(self.import_file)

        try:
            file_content = fileobj.read()
            workbook = load_workbook(filename=io.BytesIO(file_content), data_only=True)
        except Exception:
            raise UserError("Only .xlsx Excel files are supported.")

        sheet = workbook.active  # first sheet

        input_types_row = 2  # Excel rows start from 1
        employee_col = 2  # Column B
        first_row_data = 3  # Data starts from row 3

        # Last row with data in column A
        last_row_data = sheet.max_row

        # 🔹 Get input type names (row 2, starting from column C)
        input_type_names = []
        col = 3  # column C
        while sheet.cell(row=input_types_row, column=col).value:
            input_type_names.append(
                str(sheet.cell(row=input_types_row, column=col).value).strip()
            )
            col += 1

        # 🔹 Get employee names (column B, starting from row 3)
        employee_names = []
        for row in range(first_row_data, last_row_data + 1):
            employee_name = sheet.cell(row=row, column=employee_col).value
            if employee_name:
                employee_names.append(str(employee_name).strip())

        employee_data = self.validate_employees(employee_names)
        input_types = self.validate_input_types(input_type_names)

        # finish import after validation if there are problems with data
        if self.alert:
            return self.finish_import()

        payslip_vals = []

        for row in range(first_row_data, last_row_data + 1):
            employee_name = sheet.cell(row=row, column=employee_col).value

            if not employee_name:
                continue

            employee = employee_data.get(str(employee_name).strip())
            row_dict = {}

            values = []
            for idx, input_type_id in enumerate(input_types.ids):
                value = sheet.cell(row=row, column=3 + idx).value or 0
                values.append((input_type_id, value))

            row_dict[employee] = values
            payslip_vals.append(row_dict)

        self.create_payslip(payslip_vals)
        return self.finish_import()
